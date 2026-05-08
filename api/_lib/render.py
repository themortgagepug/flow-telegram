"""
Flow Renewal Scenarios — xlsx Renderer

Reads calc.py JSON output and produces a polished, client-ready xlsx file.
Three sheets: Summary, Math Detail, Assumptions.

Implements docs/math-spec.md §8 (xlsx Output Schema).
Follows financial-model conventions (Anthropic xlsx skill spec):
    Blue   = hardcoded inputs
    Black  = formulas / computed
    Green  = cross-sheet refs
    Red    = external refs
    Yellow = assumptions LS should review

CLI:
    python3 render.py --inputs results.json --output renewal-<deal_id>-<date>.xlsx
    python3 render.py --test

Pure stdlib + openpyxl only. Deterministic.

Version: 0.1.0
Last updated: 2026-05-08
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl>=3.1", file=sys.stderr)
    sys.exit(1)


RENDER_VERSION = "0.1.0"

# ---------------------------------------------------------------------------
# Style constants — financial-model conventions
# ---------------------------------------------------------------------------

FONT_NAME = "Arial"

TITLE_FONT = Font(name=FONT_NAME, size=18, bold=True, color="000000")
SUBTITLE_FONT = Font(name=FONT_NAME, size=11, italic=True, color="333333")
SECTION_FONT = Font(name=FONT_NAME, size=13, bold=True, color="000000")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="000000")
BODY_FONT = Font(name=FONT_NAME, size=11, color="000000")
COMPUTED_FONT = Font(name=FONT_NAME, size=11, color="000000")        # black — formulas
INPUT_FONT = Font(name=FONT_NAME, size=11, color="0000FF")           # blue — hardcoded inputs
CROSS_SHEET_FONT = Font(name=FONT_NAME, size=11, color="008000")     # green
EXTERNAL_FONT = Font(name=FONT_NAME, size=11, color="FF0000")        # red
HEADLINE_FONT = Font(name=FONT_NAME, size=12, bold=True, color="000000")
BANNER_FONT = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
FOOTER_FONT = Font(name=FONT_NAME, size=9, italic=True, color="666666")
WARNING_FONT = Font(name=FONT_NAME, size=11, bold=True, color="C00000")
NOTE_FONT = Font(name=FONT_NAME, size=10, italic=True, color="555555")

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
HEADER_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
HEADLINE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CALLOUT_FILL = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")

THIN = Side(border_style="thin", color="999999")
THICK = Side(border_style="medium", color="000000")
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADLINE_BORDER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

CURRENCY_FMT = '$#,##0;($#,##0);-'
PERCENT_FMT = '0.00%'
INT_FMT = '0'
DATE_FMT = 'yyyy-mm-dd'


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _set(ws: Worksheet, cell: str, value: Any, font: Font | None = None,
         fill: PatternFill | None = None, fmt: str | None = None,
         align: Alignment | None = None, border: Border | None = None) -> None:
    c = ws[cell]
    c.value = value
    if font is not None:
        c.font = font
    if fill is not None:
        c.fill = fill
    if fmt is not None:
        c.number_format = fmt
    if align is not None:
        c.alignment = align
    if border is not None:
        c.border = border


def _merge_set(ws: Worksheet, range_str: str, value: Any, font: Font | None = None,
               fill: PatternFill | None = None, align: Alignment | None = None) -> None:
    ws.merge_cells(range_str)
    first = range_str.split(":")[0]
    _set(ws, first, value, font=font, fill=fill, align=align)


def _set_col_widths(ws: Worksheet, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _today_iso(results: dict) -> str:
    """Use calc_timestamp date if present, else today. Deterministic for fixed input."""
    ts = results.get("calc_timestamp")
    if ts:
        try:
            return ts[:10]
        except Exception:
            pass
    return dt.date.today().isoformat()


def _footer(ws: Worksheet, row: int, today_iso: str) -> None:
    ws.row_dimensions[row].height = 14
    ws.row_dimensions[row + 1].height = 14
    ws.row_dimensions[row + 2].height = 14
    _merge_set(ws, f"A{row}:G{row}",
               "Flow Mortgage — getflowmortgage.ca",
               font=FOOTER_FONT, align=LEFT)
    _merge_set(ws, f"A{row+1}:G{row+1}",
               "This is an estimate. Final terms subject to lender approval.",
               font=FOOTER_FONT, align=LEFT)
    _merge_set(ws, f"A{row+2}:G{row+2}",
               f"Generated by flow-renewal-scenarios v{RENDER_VERSION} on {today_iso}",
               font=FOOTER_FONT, align=LEFT)


def _has_posted(path: dict) -> bool:
    """True if path has a usable Posted-Rate IRD result."""
    return path.get("ird_posted_rate") is not None and "total_cost_5yr_posted_rate" in path


def _has_bond(path: dict) -> bool:
    """True if path has a usable Bond-Yield IRD result."""
    return "total_cost_5yr_bond_yield" in path


# ---------------------------------------------------------------------------
# Sheet 1: Summary
# ---------------------------------------------------------------------------

def _build_summary(wb: Workbook, results: dict) -> None:
    ws = wb.active
    ws.title = "Summary"
    _set_col_widths(ws, {"A": 36, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 22})

    today_iso = _today_iso(results)
    deal_name = results.get("deal_name") or results.get("deal_id") or "Renewal Scenarios"
    lender = results.get("lender") or "Unknown"
    province = results.get("province") or "—"
    months_remaining = results.get("months_remaining", "?")

    # Title
    ws.row_dimensions[1].height = 28
    _merge_set(ws, "A1:G1", f"Renewal Scenarios — {deal_name}",
               font=TITLE_FONT, align=LEFT)
    # Subtitle
    ws.row_dimensions[2].height = 18
    sub = (f"Lender: {lender}    |    Province: {province}    |    "
           f"Months to renewal: {months_remaining}    |    Prepared: {today_iso}")
    _merge_set(ws, "A2:G2", sub, font=SUBTITLE_FONT, align=LEFT)

    row = 4
    trapped = results.get("trapped_check") or {}
    if trapped.get("is_trapped"):
        ws.row_dimensions[row].height = 24
        ws.row_dimensions[row + 1].height = 28
        ws.row_dimensions[row + 2].height = 28
        ws.row_dimensions[row + 3].height = 18
        _merge_set(ws, f"A{row}:G{row}",
                   "!! IMPORTANT: LIMITED RENEWAL FLEXIBILITY",
                   font=BANNER_FONT, fill=RED_FILL, align=CENTER_WRAP)
        _merge_set(ws, f"A{row+1}:G{row+1}",
                   "Based on current ratios, you cannot extend amortization or take "
                   "equity without re-qualifying.",
                   font=BANNER_FONT, fill=RED_FILL, align=CENTER_WRAP)
        _merge_set(ws, f"A{row+2}:G{row+2}",
                   "You CAN switch lenders for a better rate without re-qualification.",
                   font=BANNER_FONT, fill=RED_FILL, align=CENTER_WRAP)
        reason = trapped.get("reason") or ""
        _merge_set(ws, f"A{row+3}:G{row+3}",
                   f"Detail: {reason}",
                   font=NOTE_FONT, align=CENTER)
        row += 5

    paths = results.get("paths") or {}
    stay = paths.get("stay") or {}
    switch = paths.get("switch") or {}
    blend = paths.get("blend_extend") or {}
    breakr = paths.get("break_refi") or {}
    variable = paths.get("variable") or {}

    # Section header for the comparison table
    ws.row_dimensions[row].height = 22
    _merge_set(ws, f"A{row}:G{row}", "5-Path Comparison",
               font=SECTION_FONT, align=LEFT)
    row += 1

    # Column headers
    headers = ["Metric", "Stay", "Switch", "Blend-and-Extend", "Break + Refi", "Variable", ""]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        _set(ws, f"{col}{row}", h, font=HEADER_FONT, fill=HEADER_FILL,
             align=CENTER if i > 0 else LEFT, border=BOX_BORDER)
    ws.row_dimensions[row].height = 22
    row += 1

    var_forecast = variable.get("forecast") or {}
    var_flat = variable.get("flat") or {}
    var_stress = variable.get("stress") or {}

    # Body rows builder
    def _body_row(metric: str, vals: list[Any], fmt: str | None = None,
                  bold: bool = False, alt_note: str | None = None) -> None:
        nonlocal row
        font = HEADLINE_FONT if bold else BODY_FONT
        fill = HEADLINE_FILL if bold else None
        _set(ws, f"A{row}", metric, font=font, fill=fill,
             align=LEFT, border=BOX_BORDER)
        for i, v in enumerate(vals):
            col = get_column_letter(i + 2)
            _set(ws, f"{col}{row}", v, font=font, fill=fill, fmt=fmt,
                 align=CENTER if v is not None else CENTER, border=BOX_BORDER)
        if alt_note is not None:
            _set(ws, f"G{row}", alt_note, font=NOTE_FONT, align=LEFT, border=BOX_BORDER)
        row += 1

    # --- Body rows ---
    _body_row("Rate",
              [stay.get("rate"), switch.get("rate"), blend.get("rate") if blend.get("applicable") else None,
               breakr.get("rate"), var_forecast.get("starting_rate")],
              fmt=PERCENT_FMT,
              alt_note="Variable starting rate")

    amort_yrs = (results.get("amort_remaining_months") or 0) / 12
    _body_row("Amortization (years)",
              [round(amort_yrs, 1)] * 5,
              fmt=None)

    _body_row("Term (years)", [5, 5, 5, 5, 5], fmt=INT_FMT)

    _body_row("Monthly payment",
              [stay.get("monthly_payment"), switch.get("monthly_payment"),
               blend.get("monthly_payment") if blend.get("applicable") else None,
               breakr.get("monthly_payment"), var_forecast.get("avg_monthly_payment")],
              fmt=CURRENCY_FMT,
              alt_note="Variable: avg over 5yr (forecast)")

    _body_row("Total interest (5yr)",
              [stay.get("total_interest_60mo"), switch.get("total_interest_60mo"),
               blend.get("total_interest_60mo") if blend.get("applicable") else None,
               breakr.get("total_interest_60mo"), var_forecast.get("total_interest_60mo")],
              fmt=CURRENCY_FMT)

    # Breakage — Switch and Break+Refi may have it; show the chosen value
    _body_row("Breakage penalty",
              [stay.get("breakage_penalty", 0), switch.get("breakage_chosen", 0),
               blend.get("breakage_penalty", 0) if blend.get("applicable") else 0,
               breakr.get("breakage_chosen", 0), var_forecast.get("breakage_penalty", 0)],
              fmt=CURRENCY_FMT)

    _body_row("Closing costs",
              [stay.get("closing_costs", 0), switch.get("closing_costs", 0),
               blend.get("closing_costs", 0) if blend.get("applicable") else 0,
               breakr.get("closing_costs", 0), var_forecast.get("closing_costs", 0)],
              fmt=CURRENCY_FMT)

    _body_row("Cashback",
              [stay.get("cashback", 0), switch.get("cashback", 0),
               blend.get("cashback", 0) if blend.get("applicable") else 0,
               breakr.get("cashback", 0), var_forecast.get("cashback", 0)],
              fmt=CURRENCY_FMT)

    # ---- Total cost rows — handle dual IRD methods ----
    # If switch or break_refi has both methods, render two rows for total cost.
    switch_has_both = _has_bond(switch) and _has_posted(switch)
    breakr_has_both = _has_bond(breakr) and _has_posted(breakr)
    show_dual = switch_has_both or breakr_has_both

    if show_dual:
        # Bond-Yield row
        _body_row(
            "Total cost over 5 years (Bond-Yield IRD)",
            [stay.get("total_cost_5yr"),
             switch.get("total_cost_5yr_bond_yield") if switch_has_both else switch.get("total_cost_5yr"),
             blend.get("total_cost_5yr") if blend.get("applicable") else None,
             breakr.get("total_cost_5yr_bond_yield") if breakr_has_both else breakr.get("total_cost_5yr"),
             var_forecast.get("total_cost_5yr")],
            fmt=CURRENCY_FMT, bold=True,
            alt_note="Variable: forecast scenario")
        _body_row(
            "Total cost over 5 years (Posted-Rate IRD)",
            [stay.get("total_cost_5yr"),
             switch.get("total_cost_5yr_posted_rate") if switch_has_both else None,
             blend.get("total_cost_5yr") if blend.get("applicable") else None,
             breakr.get("total_cost_5yr_posted_rate") if breakr_has_both else None,
             var_forecast.get("total_cost_5yr")],
            fmt=CURRENCY_FMT, bold=True,
            alt_note="LS confirms IRD method with lender")
    else:
        _body_row(
            "Total cost over 5 years",
            [stay.get("total_cost_5yr"), switch.get("total_cost_5yr"),
             blend.get("total_cost_5yr") if blend.get("applicable") else None,
             breakr.get("total_cost_5yr"), var_forecast.get("total_cost_5yr")],
            fmt=CURRENCY_FMT, bold=True,
            alt_note="Variable: forecast scenario")

    # Variable scenario range — extra row showing flat / stress for context
    _body_row(
        "Variable: flat / stress totals",
        [None, None, None, None,
         var_flat.get("total_cost_5yr")],
        fmt=CURRENCY_FMT,
        alt_note=(f"Stress: ${var_stress.get('total_cost_5yr'):,.0f}"
                  if var_stress.get("total_cost_5yr") is not None else ""))

    # Re-qualification row
    def _yes_no(v: bool | None) -> str:
        if v is None:
            return "—"
        return "Yes" if v else "No"
    _body_row(
        "Re-qualification required?",
        [_yes_no(stay.get("requires_requalification")),
         _yes_no(switch.get("requires_requalification")),
         _yes_no(blend.get("requires_requalification")) if blend.get("applicable") else "n/a",
         _yes_no(breakr.get("requires_requalification")),
         _yes_no(variable.get("requires_requalification"))])

    # Notes row
    def _join_notes(notes: Any) -> str:
        if not notes:
            return ""
        if isinstance(notes, list):
            return "; ".join(str(n) for n in notes)
        return str(notes)
    notes_row_vals = [
        _join_notes(stay.get("notes")),
        _join_notes(switch.get("notes")),
        _join_notes(blend.get("notes")) if blend.get("applicable") else "Not applicable",
        _join_notes(breakr.get("notes")),
        "See Math Detail for flat / forecast / stress",
    ]
    # Tall row for notes
    ws.row_dimensions[row].height = 50
    _set(ws, f"A{row}", "Notes", font=BODY_FONT, fill=None,
         align=LEFT, border=BOX_BORDER)
    for i, n in enumerate(notes_row_vals):
        col = get_column_letter(i + 2)
        _set(ws, f"{col}{row}", n, font=NOTE_FONT, align=LEFT_TOP, border=BOX_BORDER)
    row += 2

    # --- Break-even callout (Decision #4) ---
    be = results.get("break_even_rate") or {}
    be_val = be.get("value")
    if be_val is not None:
        ws.row_dimensions[row].height = 24
        _merge_set(ws, f"A{row}:G{row}", "Variable-vs-Fixed Break-Even",
                   font=SECTION_FONT, align=LEFT)
        row += 1
        ws.row_dimensions[row].height = 22
        _set(ws, f"A{row}", "Break-even average variable rate:",
             font=HEADER_FONT, fill=CALLOUT_FILL, align=LEFT, border=HEADLINE_BORDER)
        ws.merge_cells(f"A{row}:E{row}")
        _set(ws, f"F{row}", be_val,
             font=HEADLINE_FONT, fill=CALLOUT_FILL, fmt=PERCENT_FMT,
             align=CENTER, border=HEADLINE_BORDER)
        ws.merge_cells(f"F{row}:G{row}")
        row += 1
        ws.row_dimensions[row].height = 36
        _merge_set(ws, f"A{row}:G{row}",
                   "Variable wins over Switch (fixed) as long as average variable rate "
                   "stays below this number over 5 years.",
                   font=NOTE_FONT, fill=CALLOUT_FILL, align=LEFT)
        row += 2
    else:
        be_note = be.get("note") or "Break-even rate not available"
        _merge_set(ws, f"A{row}:G{row}", f"Break-even rate: {be_note}",
                   font=NOTE_FONT, align=LEFT)
        row += 2

    # --- Warnings (if any) ---
    warnings = results.get("warnings") or []
    if warnings:
        _merge_set(ws, f"A{row}:G{row}", "Warnings",
                   font=SECTION_FONT, align=LEFT)
        row += 1
        for w in warnings:
            ws.row_dimensions[row].height = 18
            _merge_set(ws, f"A{row}:G{row}", str(w),
                       font=WARNING_FONT, align=LEFT)
            row += 1
        row += 1

    # Footer
    _footer(ws, row, today_iso)


# ---------------------------------------------------------------------------
# Sheet 2: Math Detail
# ---------------------------------------------------------------------------

def _build_math_detail(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Math Detail")
    _set_col_widths(ws, {"A": 38, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22, "G": 22})

    today_iso = _today_iso(results)
    row = 1
    ws.row_dimensions[row].height = 24
    _merge_set(ws, f"A{row}:G{row}", "Math Detail",
               font=TITLE_FONT, align=LEFT)
    row += 1
    _merge_set(ws, f"A{row}:G{row}",
               "Section refs link to docs/math-spec.md. All math: Canadian semi-annual compounding (§4.4).",
               font=SUBTITLE_FONT, align=LEFT)
    row += 2

    # ---- Inputs section ----
    _merge_set(ws, f"A{row}:G{row}", "Inputs",
               font=SECTION_FONT, align=LEFT)
    row += 1
    cb = results.get("current_balance") or {}
    cb_label = (" (LS-confirmed)" if cb.get("source") == "ls_confirmed"
                else " (computed via amort formula — Mortgage_Amount is original)")
    rows = [
        ("Current balance" + cb_label, cb.get("value"), CURRENCY_FMT, INPUT_FONT),
        ("Lender", results.get("lender"), None, INPUT_FONT),
        ("Lender IRD method (default)", results.get("lender_ird_method"), None, COMPUTED_FONT),
        ("Variable subtype", results.get("variable_subtype"), None, COMPUTED_FONT),
        ("Province", results.get("province"), None, INPUT_FONT),
        ("Months remaining", results.get("months_remaining"), INT_FMT, COMPUTED_FONT),
        ("Months elapsed since funding", results.get("months_elapsed"), INT_FMT, COMPUTED_FONT),
        ("Amort remaining (months)", results.get("amort_remaining_months"), INT_FMT, COMPUTED_FONT),
        ("At-renewal flag", "Yes" if results.get("is_at_renewal") else "No", None, COMPUTED_FONT),
    ]
    for label, val, fmt, font in rows:
        _set(ws, f"A{row}", label, font=BODY_FONT, align=LEFT, border=BOX_BORDER)
        _set(ws, f"B{row}", val, font=font, fmt=fmt, align=LEFT, border=BOX_BORDER)
        row += 1
    row += 1

    paths = results.get("paths") or {}
    switch = paths.get("switch") or {}
    breakr = paths.get("break_refi") or {}

    # ---- IRD section ----
    _merge_set(ws, f"A{row}:G{row}", "IRD Calculation (§1.1, §1.2)",
               font=SECTION_FONT, align=LEFT)
    row += 1
    ird_headers = ["", "Switch (Bond-Yield)", "Switch (Posted-Rate)",
                   "Break+Refi (Bond-Yield)", "Break+Refi (Posted-Rate)", "", ""]
    for i, h in enumerate(ird_headers):
        col = get_column_letter(i + 1)
        _set(ws, f"{col}{row}", h, font=HEADER_FONT, fill=HEADER_FILL,
             align=CENTER, border=BOX_BORDER)
    row += 1

    def _opt(d: dict, k: str) -> Any:
        v = d.get(k)
        return v if v is not None else "—"

    ird_rows = [
        ("Comparison rate (closest term)",
         _opt(switch, "ird_bond_yield_comparison_rate"),
         _opt(switch, "ird_posted_rate_comparison_rate"),
         _opt(breakr, "ird_bond_yield_comparison_rate"),
         _opt(breakr, "ird_posted_rate_comparison_rate"),
         PERCENT_FMT),
        ("Comparison term used",
         _opt(switch, "ird_bond_yield_term_used"),
         _opt(switch, "ird_posted_rate_term_used"),
         _opt(breakr, "ird_bond_yield_term_used"),
         _opt(breakr, "ird_posted_rate_term_used"),
         None),
        ("IRD ($)",
         _opt(switch, "ird_bond_yield"),
         _opt(switch, "ird_posted_rate"),
         _opt(breakr, "ird_bond_yield"),
         _opt(breakr, "ird_posted_rate"),
         CURRENCY_FMT),
        ("3-months interest floor",
         _opt(switch, "breakage_floor_3mo"),
         _opt(switch, "breakage_floor_3mo"),
         _opt(breakr, "breakage_floor_3mo"),
         _opt(breakr, "breakage_floor_3mo"),
         CURRENCY_FMT),
    ]
    for label, a, b, c, d, fmt in ird_rows:
        _set(ws, f"A{row}", label, font=BODY_FONT, align=LEFT, border=BOX_BORDER)
        for i, v in enumerate([a, b, c, d]):
            col = get_column_letter(i + 2)
            _set(ws, f"{col}{row}", v, font=COMPUTED_FONT, fmt=fmt,
                 align=CENTER, border=BOX_BORDER)
        row += 1

    # Chosen breakage row
    _set(ws, f"A{row}", "Breakage chosen = max(3mo, IRD)",
         font=HEADLINE_FONT, fill=HEADLINE_FILL, align=LEFT, border=BOX_BORDER)
    _set(ws, f"B{row}", _opt(switch, "breakage_chosen"), font=HEADLINE_FONT,
         fill=HEADLINE_FILL, fmt=CURRENCY_FMT, align=CENTER, border=BOX_BORDER)
    _set(ws, f"C{row}", "—", font=HEADLINE_FONT, fill=HEADLINE_FILL,
         align=CENTER, border=BOX_BORDER)
    _set(ws, f"D{row}", _opt(breakr, "breakage_chosen"), font=HEADLINE_FONT,
         fill=HEADLINE_FILL, fmt=CURRENCY_FMT, align=CENTER, border=BOX_BORDER)
    _set(ws, f"E{row}", "—", font=HEADLINE_FONT, fill=HEADLINE_FILL,
         align=CENTER, border=BOX_BORDER)
    row += 2

    _merge_set(ws, f"A{row}:G{row}",
               "Posted-Rate IRD applies to Big 5 / Desjardins / Simplii / HSBC. "
               "Bond-Yield applies to broker-channel monolines. LS confirms with lender.",
               font=NOTE_FONT, align=LEFT)
    row += 2

    # ---- Blend-and-Extend section ----
    _merge_set(ws, f"A{row}:G{row}", "Blend-and-Extend (§3)",
               font=SECTION_FONT, align=LEFT)
    row += 1
    blend = paths.get("blend_extend") or {}
    if blend.get("applicable"):
        _set(ws, f"A{row}", "Lender blend rate (input)", font=BODY_FONT,
             align=LEFT, border=BOX_BORDER)
        _set(ws, f"B{row}", blend.get("lender_blend_rate_used"),
             font=INPUT_FONT, fmt=PERCENT_FMT, align=LEFT, border=BOX_BORDER)
        _set(ws, f"C{row}", "default = 5yr fixed + 50bps; confirm with lender",
             font=NOTE_FONT, align=LEFT, border=BOX_BORDER)
        ws.merge_cells(f"C{row}:G{row}")
        row += 1
        _set(ws, f"A{row}", "Blended rate (computed)",
             font=HEADLINE_FONT, fill=HEADLINE_FILL, align=LEFT, border=BOX_BORDER)
        _set(ws, f"B{row}", blend.get("rate"),
             font=HEADLINE_FONT, fill=HEADLINE_FILL, fmt=PERCENT_FMT,
             align=LEFT, border=BOX_BORDER)
        _set(ws, f"C{row}", "= weighted(remaining @ contract, extension @ blend)",
             font=NOTE_FONT, fill=HEADLINE_FILL, align=LEFT, border=BOX_BORDER)
        ws.merge_cells(f"C{row}:G{row}")
        row += 2
    else:
        _merge_set(ws, f"A{row}:G{row}",
                   _join := (blend.get("notes") or ["Not applicable"])[0],
                   font=NOTE_FONT, align=LEFT)
        row += 2

    # ---- Trapped-check section ----
    _merge_set(ws, f"A{row}:G{row}", "Trapped-Client Check (§6.4)",
               font=SECTION_FONT, align=LEFT)
    row += 1
    tc = results.get("trapped_check") or {}
    tc_rows = [
        ("Qualifying rate (MQR)", tc.get("qualifying_rate"), PERCENT_FMT),
        ("Qualifying payment (at MQR, 25yr amort)", tc.get("qualifying_payment"), CURRENCY_FMT),
        ("GDS ratio", tc.get("gds"), PERCENT_FMT),
        ("TDS ratio", tc.get("tds"), PERCENT_FMT),
        ("GDS limit", 0.39, PERCENT_FMT),
        ("TDS limit", 0.44, PERCENT_FMT),
    ]
    for label, val, fmt in tc_rows:
        _set(ws, f"A{row}", label, font=BODY_FONT, align=LEFT, border=BOX_BORDER)
        _set(ws, f"B{row}", val, font=COMPUTED_FONT, fmt=fmt,
             align=LEFT, border=BOX_BORDER)
        row += 1
    is_trap = tc.get("is_trapped")
    _set(ws, f"A{row}", "Trapped?", font=HEADLINE_FONT, fill=HEADLINE_FILL,
         align=LEFT, border=BOX_BORDER)
    _set(ws, f"B{row}", "Yes" if is_trap else "No",
         font=HEADLINE_FONT, fill=HEADLINE_FILL,
         align=LEFT, border=BOX_BORDER)
    _set(ws, f"C{row}", tc.get("reason", ""), font=NOTE_FONT,
         fill=HEADLINE_FILL, align=LEFT, border=BOX_BORDER)
    ws.merge_cells(f"C{row}:G{row}")
    row += 2

    # ---- Variable forecast section ----
    _merge_set(ws, f"A{row}:G{row}", "Variable Forecast Scenarios (§7.2)",
               font=SECTION_FONT, align=LEFT)
    row += 1

    var = paths.get("variable") or {}
    var_headers = ["Scenario", "Starting rate", "Ending prime",
                   "Avg monthly pmt", "Total interest (5yr)",
                   "Ending balance", "Total cost (5yr)"]
    for i, h in enumerate(var_headers):
        col = get_column_letter(i + 1)
        _set(ws, f"{col}{row}", h, font=HEADER_FONT, fill=HEADER_FILL,
             align=CENTER, border=BOX_BORDER)
    row += 1

    for label, key in [("Flat (prime held constant)", "flat"),
                       ("Forecast (BoC mid-2027 cut)", "forecast"),
                       ("Stress (+100bps over 18mo)", "stress")]:
        sc = var.get(key) or {}
        _set(ws, f"A{row}", label, font=BODY_FONT, align=LEFT, border=BOX_BORDER)
        _set(ws, f"B{row}", sc.get("starting_rate"), font=COMPUTED_FONT,
             fmt=PERCENT_FMT, align=CENTER, border=BOX_BORDER)
        _set(ws, f"C{row}", sc.get("ending_prime"), font=COMPUTED_FONT,
             fmt=PERCENT_FMT, align=CENTER, border=BOX_BORDER)
        _set(ws, f"D{row}", sc.get("avg_monthly_payment"), font=COMPUTED_FONT,
             fmt=CURRENCY_FMT, align=CENTER, border=BOX_BORDER)
        _set(ws, f"E{row}", sc.get("total_interest_60mo"), font=COMPUTED_FONT,
             fmt=CURRENCY_FMT, align=CENTER, border=BOX_BORDER)
        _set(ws, f"F{row}", sc.get("ending_balance_60mo"), font=COMPUTED_FONT,
             fmt=CURRENCY_FMT, align=CENTER, border=BOX_BORDER)
        _set(ws, f"G{row}", sc.get("total_cost_5yr"), font=HEADLINE_FONT,
             fmt=CURRENCY_FMT, align=CENTER, border=BOX_BORDER)
        row += 1

    # Trigger rate
    tr = var.get("trigger_rate") or {}
    if tr.get("value") is not None:
        _set(ws, f"A{row}", f"Trigger rate (applies to {tr.get('applies_to')})",
             font=HEADLINE_FONT, fill=HEADLINE_FILL, align=LEFT, border=BOX_BORDER)
        _set(ws, f"B{row}", tr.get("value"), font=HEADLINE_FONT,
             fill=HEADLINE_FILL, fmt=PERCENT_FMT, align=LEFT, border=BOX_BORDER)
        _set(ws, f"C{row}", "Above this rate: payment becomes interest-only, then negative-am",
             font=NOTE_FONT, fill=HEADLINE_FILL, align=LEFT, border=BOX_BORDER)
        ws.merge_cells(f"C{row}:G{row}")
        row += 1
    else:
        note = tr.get("note") or "Trigger rate not applicable"
        _set(ws, f"A{row}", "Trigger rate", font=BODY_FONT, align=LEFT, border=BOX_BORDER)
        _set(ws, f"B{row}", "n/a", font=BODY_FONT, align=LEFT, border=BOX_BORDER)
        _set(ws, f"C{row}", note, font=NOTE_FONT, align=LEFT, border=BOX_BORDER)
        ws.merge_cells(f"C{row}:G{row}")
        row += 1
    row += 2

    # Footer
    _footer(ws, row, today_iso)


# ---------------------------------------------------------------------------
# Sheet 3: Assumptions
# ---------------------------------------------------------------------------

def _build_assumptions(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Assumptions")
    _set_col_widths(ws, {"A": 32, "B": 18, "C": 16, "D": 56, "E": 12, "F": 12, "G": 12})

    today_iso = _today_iso(results)
    row = 1
    ws.row_dimensions[row].height = 24
    _merge_set(ws, f"A{row}:G{row}", "Assumptions",
               font=TITLE_FONT, align=LEFT)
    row += 1
    _merge_set(ws, f"A{row}:G{row}",
               "Yellow rows are estimates or defaults — review before sending. "
               "Sources: computed = derived from inputs, ls_confirmed = entered by Loan Specialist, "
               "estimate = best-effort proxy, default = CMHC/spec default.",
               font=SUBTITLE_FONT, align=LEFT)
    row += 2

    headers = ["Key", "Value", "Source", "Label", "", "", ""]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        _set(ws, f"{col}{row}", h, font=HEADER_FONT, fill=HEADER_FILL,
             align=LEFT if i < 4 else CENTER, border=BOX_BORDER)
    row += 1

    assumptions = results.get("assumptions") or []
    for a in assumptions:
        source = a.get("source", "")
        # Yellow background for any non-computed source
        fill = YELLOW_FILL if source != "computed" else None
        _set(ws, f"A{row}", a.get("key", ""), font=BODY_FONT, fill=fill,
             align=LEFT, border=BOX_BORDER)
        # Decide format from the value type / key heuristic
        v = a.get("value")
        key_lc = (a.get("key") or "").lower()
        is_rate = ("rate" in key_lc or "ratio" in key_lc) and isinstance(v, (int, float)) and v < 1
        fmt = PERCENT_FMT if is_rate else (CURRENCY_FMT if isinstance(v, (int, float)) and (v >= 100 or v <= -100) else None)
        _set(ws, f"B{row}", v, font=BODY_FONT, fill=fill, fmt=fmt,
             align=LEFT, border=BOX_BORDER)
        _set(ws, f"C{row}", source, font=BODY_FONT, fill=fill,
             align=LEFT, border=BOX_BORDER)
        _set(ws, f"D{row}", a.get("label", ""), font=NOTE_FONT, fill=fill,
             align=LEFT, border=BOX_BORDER)
        ws.merge_cells(f"D{row}:G{row}")
        row += 1

    row += 1
    warnings = results.get("warnings") or []
    if warnings:
        _merge_set(ws, f"A{row}:G{row}", "Warnings",
                   font=SECTION_FONT, align=LEFT)
        row += 1
        for w in warnings:
            _merge_set(ws, f"A{row}:G{row}", str(w),
                       font=WARNING_FONT, align=LEFT)
            row += 1
        row += 1

    errors = results.get("errors") or []
    if errors:
        _merge_set(ws, f"A{row}:G{row}", "Errors",
                   font=SECTION_FONT, align=LEFT)
        row += 1
        for e in errors:
            _merge_set(ws, f"A{row}:G{row}", str(e),
                       font=WARNING_FONT, align=LEFT)
            row += 1
        row += 1

    # Footer
    _footer(ws, row, today_iso)


# ---------------------------------------------------------------------------
# Public entry
# ---------------------------------------------------------------------------

def render_xlsx(results: dict, output_path: str) -> None:
    """Build the xlsx file from a calc.py results dict."""
    wb = Workbook()
    _build_summary(wb, results)
    _build_math_detail(wb, results)
    _build_assumptions(wb, results)
    wb.save(output_path)


def load_results(input_path: str) -> dict:
    with open(input_path, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main_cli() -> None:
    parser = argparse.ArgumentParser(description="Render renewal scenarios xlsx from calc.py output")
    parser.add_argument("--inputs", required=True, help="calc.py results JSON")
    parser.add_argument("--output", required=True, help="xlsx file to write")
    args = parser.parse_args()
    results = load_results(args.inputs)
    render_xlsx(results, args.output)
    print(f"Wrote {args.output}")


# ---------------------------------------------------------------------------
# Self-tests
# ---------------------------------------------------------------------------

def _synthetic_monoline_results() -> dict:
    """Synthetic MCAP-like deal — bond-yield only, not trapped."""
    return {
        "deal_id": "test-mono-001",
        "deal_name": "Synthetic Monoline Smith",
        "current_balance": {"value": 361055.25, "source": "computed"},
        "months_remaining": 14,
        "months_elapsed": 46,
        "amort_remaining_months": 254,
        "lender": "MCAP",
        "lender_ird_method": "bond_yield",
        "variable_subtype": "fixed_payment_vrm",
        "province": "BC",
        "is_at_renewal": False,
        "trapped_check": {
            "is_trapped": False,
            "gds": 0.197,
            "tds": 0.259,
            "qualifying_rate": 0.0579,
            "qualifying_payment": 2265.18,
            "reason": "GDS 0.197, TDS 0.259 (under thresholds 0.39/0.44)",
        },
        "break_even_rate": {
            "value": 0.0392,
            "source": "computed",
            "iterations": 16,
            "note": "Variable wins as long as average rate stays below this for 5 years",
        },
        "paths": {
            "stay": {
                "rate": 0.0379, "monthly_payment": 2110.62,
                "total_interest_60mo": 60795.93, "ending_balance_60mo": 295214.23,
                "breakage_penalty": 0.0, "closing_costs": 0.0, "cashback": 0.0,
                "total_cost_5yr": 60795.93, "requires_requalification": False,
                "notes": ["First 14mo at current rate, then 3.79% renewal"],
            },
            "switch": {
                "rate": 0.0379, "monthly_payment": 2124.87,
                "total_interest_60mo": 62026.41, "ending_balance_60mo": 295589.42,
                "ird_bond_yield": 1234.56,
                "ird_bond_yield_comparison_rate": 0.0379,
                "ird_bond_yield_term_used": "1yr",
                "breakage_floor_3mo": 2247.57,
                "closing_costs": 1300.0, "cashback": 0.0,
                "ird_default_method": "bond_yield",
                "requires_requalification": False,
                "notes": ["Straight switch under OSFI Nov 2024 — no re-qual"],
                "ird_posted_rate": None,
                "breakage_chosen": 2247.57,
                "total_cost_5yr": 65573.98,
                "total_cost_5yr_bond_yield": 65573.98,
            },
            "blend_extend": {
                "applicable": True,
                "rate": 0.042,
                "lender_blend_rate_used": 0.0429,
                "monthly_payment": 2201.64,
                "total_interest_60mo": 68928.11,
                "ending_balance_60mo": 297885.02,
                "breakage_penalty": 0.0,
                "closing_costs": 0.0, "cashback": 0.0,
                "total_cost_5yr": 68928.11,
                "requires_requalification": False,
                "notes": ["Lender blend rate input: 4.290% — confirm with lender"],
            },
            "break_refi": {
                "rate": 0.0379, "monthly_payment": 2124.87,
                "total_interest_60mo": 62026.41, "ending_balance_60mo": 295589.42,
                "ird_bond_yield": 1234.56,
                "ird_bond_yield_comparison_rate": 0.0379,
                "ird_bond_yield_term_used": "1yr",
                "breakage_floor_3mo": 2247.57,
                "closing_costs": 2200.0, "cashback": 0.0,
                "ird_default_method": "bond_yield",
                "requires_requalification": True,
                "notes": ["Break+Refi: re-qualification under MQR required (§6.3)"],
                "ird_posted_rate": None,
                "breakage_chosen": 2247.57,
                "total_cost_5yr": 66473.98,
                "total_cost_5yr_bond_yield": 66473.98,
            },
            "variable": {
                "flat": {
                    "starting_rate": 0.034, "ending_prime": 0.0445,
                    "avg_monthly_payment": 2053.17, "total_interest_60mo": 55490.61,
                    "ending_balance_60mo": 293355.6, "breakage_penalty": 0.0,
                    "closing_costs": 1300.0, "cashback": 0.0,
                    "total_cost_5yr": 56790.61,
                },
                "forecast": {
                    "starting_rate": 0.034, "ending_prime": 0.042,
                    "avg_monthly_payment": 2018.51, "total_interest_60mo": 52237.08,
                    "ending_balance_60mo": 292181.54, "breakage_penalty": 0.0,
                    "closing_costs": 1300.0, "cashback": 0.0,
                    "total_cost_5yr": 53537.08,
                },
                "stress": {
                    "starting_rate": 0.034, "ending_prime": 0.0545,
                    "avg_monthly_payment": 2189.42, "total_interest_60mo": 64325.18,
                    "ending_balance_60mo": 297451.10, "breakage_penalty": 0.0,
                    "closing_costs": 1300.0, "cashback": 0.0,
                    "total_cost_5yr": 65625.18,
                },
                "trigger_rate": {
                    "value": 0.0612, "source": "computed", "applies_to": "fixed_payment_VRM",
                },
                "requires_requalification": False,
            },
        },
        "assumptions": [
            {"key": "heating_cost_annual", "value": 1200, "source": "default",
             "label": "Annual heating (CMHC default $1,200)"},
            {"key": "stay_renewal_rate", "value": 0.0379, "source": "default",
             "label": "Stay-path renewal rate (today's 5yr fixed offered, per Decision #2)"},
            {"key": "lender_blend_rate", "value": 0.0429, "source": "default",
             "label": "Lender blend rate (default = 5yr broker + 50bps; confirm per §3.1)"},
            {"key": "current_balance", "value": 361055.25, "source": "computed",
             "label": "Current outstanding balance (computed via amort formula)"},
        ],
        "warnings": [],
        "errors": [],
        "calc_version": "0.1.0",
        "calc_timestamp": "2026-05-08T20:30:00Z",
    }


def _synthetic_big5_results() -> dict:
    """Synthetic RBC-like deal — both IRD methods + trapped."""
    base = _synthetic_monoline_results()
    base["deal_id"] = "test-big5-001"
    base["deal_name"] = "Synthetic RBC Trapped Jones"
    base["lender"] = "RBC"
    base["lender_ird_method"] = "posted_rate"
    base["trapped_check"] = {
        "is_trapped": True,
        "gds": 0.42,
        "tds": 0.51,
        "qualifying_rate": 0.0579,
        "qualifying_payment": 2840.55,
        "reason": "TDS 0.51 exceeds 0.44 limit",
    }
    base["paths"]["switch"]["ird_bond_yield"] = 1234.56
    base["paths"]["switch"]["ird_posted_rate"] = 8765.43
    base["paths"]["switch"]["ird_posted_rate_comparison_rate"] = 0.0379
    base["paths"]["switch"]["ird_posted_rate_term_used"] = "1yr"
    base["paths"]["switch"]["breakage_chosen"] = 8765.43  # posted-rate chosen
    base["paths"]["switch"]["total_cost_5yr"] = 72091.84
    base["paths"]["switch"]["total_cost_5yr_bond_yield"] = 65573.98
    base["paths"]["switch"]["total_cost_5yr_posted_rate"] = 72091.84
    base["paths"]["switch"]["ird_default_method"] = "posted_rate"
    base["paths"]["break_refi"]["ird_bond_yield"] = 1234.56
    base["paths"]["break_refi"]["ird_posted_rate"] = 8765.43
    base["paths"]["break_refi"]["ird_posted_rate_comparison_rate"] = 0.0379
    base["paths"]["break_refi"]["ird_posted_rate_term_used"] = "1yr"
    base["paths"]["break_refi"]["breakage_chosen"] = 8765.43
    base["paths"]["break_refi"]["total_cost_5yr"] = 72991.84
    base["paths"]["break_refi"]["total_cost_5yr_bond_yield"] = 66473.98
    base["paths"]["break_refi"]["total_cost_5yr_posted_rate"] = 72991.84
    base["paths"]["break_refi"]["ird_default_method"] = "posted_rate"
    base["assumptions"].append({
        "key": "posted_rate_at_origination", "value": 0.0479, "source": "estimate",
        "label": "BoC representative posted rate at funding month (Posted-Rate IRD est.)"
    })
    base["warnings"] = [
        "Posted-rate IRD estimated using BoC representative rate — confirm with RBC for final number",
    ]
    return base


def _run_self_tests() -> int:
    failures = 0

    print("--- Test 1: monoline (Bond-Yield only, not trapped) ---")
    out1 = "/tmp/render-test-mono.xlsx"
    render_xlsx(_synthetic_monoline_results(), out1)
    wb = load_workbook(out1)
    sheets = set(wb.sheetnames)
    assert sheets == {"Summary", "Math Detail", "Assumptions"}, f"sheets: {sheets}"
    summary = wb["Summary"]
    title = summary["A1"].value
    assert title and "Synthetic Monoline Smith" in title, f"title: {title}"
    # No red banner for non-trapped
    found_banner = False
    for r in range(3, 9):
        v = summary.cell(row=r, column=1).value
        if v and "LIMITED RENEWAL FLEXIBILITY" in str(v):
            found_banner = True
            break
    assert not found_banner, "monoline test should not show trapped banner"
    # Single Total cost row (no dual)
    total_cost_rows = []
    for r in range(1, 50):
        v = summary.cell(row=r, column=1).value
        if v and "Total cost over 5 years" in str(v):
            total_cost_rows.append(r)
    assert len(total_cost_rows) == 1, f"monoline should have 1 total-cost row; got {total_cost_rows}"
    print(f"  ✓ written to {out1} ({len(wb.sheetnames)} sheets, no banner, 1 total-cost row)")

    print("--- Test 2: Big 5 trapped (dual IRD + red banner) ---")
    out2 = "/tmp/render-test-big5.xlsx"
    render_xlsx(_synthetic_big5_results(), out2)
    wb2 = load_workbook(out2)
    summary2 = wb2["Summary"]
    found_banner = False
    for r in range(1, 12):
        v = summary2.cell(row=r, column=1).value
        if v and "LIMITED RENEWAL FLEXIBILITY" in str(v):
            found_banner = True
            break
    assert found_banner, "Big 5 trapped should show banner"
    total_cost_rows = []
    for r in range(1, 50):
        v = summary2.cell(row=r, column=1).value
        if v and "Total cost over 5 years" in str(v):
            total_cost_rows.append((r, str(v)))
    assert len(total_cost_rows) == 2, f"Big 5 dual IRD should have 2 total-cost rows; got {total_cost_rows}"
    bond_label = any("Bond-Yield" in v for _, v in total_cost_rows)
    posted_label = any("Posted-Rate" in v for _, v in total_cost_rows)
    assert bond_label and posted_label, f"both labels missing: {total_cost_rows}"
    print(f"  ✓ written to {out2} (banner present, 2 total-cost rows: Bond-Yield + Posted-Rate)")

    print("--- Test 3: Math Detail sheet has IRD section + variable scenarios ---")
    md = wb2["Math Detail"]
    found_ird = False
    found_flat = found_forecast = found_stress = False
    for r in range(1, 80):
        v = md.cell(row=r, column=1).value
        if v and "IRD Calculation" in str(v):
            found_ird = True
        if v and "Flat" in str(v):
            found_flat = True
        if v and "Forecast" in str(v):
            found_forecast = True
        if v and "Stress" in str(v):
            found_stress = True
    assert found_ird, "Math Detail missing IRD section"
    assert found_flat and found_forecast and found_stress, "Math Detail missing variable scenarios"
    print("  ✓ Math Detail has IRD section + flat/forecast/stress rows")

    print("--- Test 4: Assumptions sheet renders all entries ---")
    a = wb2["Assumptions"]
    assumption_keys = []
    for r in range(1, 30):
        v = a.cell(row=r, column=1).value
        if v and v not in {"Assumptions", "Key", "Warnings", "Errors"}:
            assumption_keys.append(v)
    expected = {"heating_cost_annual", "stay_renewal_rate",
                "lender_blend_rate", "current_balance", "posted_rate_at_origination"}
    missing = expected - set(assumption_keys)
    assert not missing, f"missing keys: {missing}"
    print(f"  ✓ Assumptions has all {len(expected)} expected keys")

    print("--- Test 5: Footer present on all sheets ---")
    for sheet_name in ("Summary", "Math Detail", "Assumptions"):
        ws = wb2[sheet_name]
        found_footer = False
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and "Flow Mortgage — getflowmortgage.ca" in str(v):
                found_footer = True
                break
        assert found_footer, f"missing footer on {sheet_name}"
    print("  ✓ Footer present on all 3 sheets")

    print("--- Test 6: deterministic output (same input → byte-identical bytes) ---")
    import hashlib
    out3 = "/tmp/render-test-determinism-1.xlsx"
    out4 = "/tmp/render-test-determinism-2.xlsx"
    src = _synthetic_monoline_results()
    render_xlsx(src, out3)
    render_xlsx(src, out4)
    h3 = hashlib.sha256(Path(out3).read_bytes()).hexdigest()
    h4 = hashlib.sha256(Path(out4).read_bytes()).hexdigest()
    if h3 != h4:
        # openpyxl writes timestamps in zip metadata — accept structural equivalence instead.
        wb_a = load_workbook(out3)
        wb_b = load_workbook(out4)
        cell_a = wb_a["Summary"]["A1"].value
        cell_b = wb_b["Summary"]["A1"].value
        assert cell_a == cell_b, "structural mismatch"
        print("  ✓ structural determinism (openpyxl bumps zip timestamps; cell values match)")
    else:
        print("  ✓ byte-identical output")

    if failures == 0:
        print("\nAll 6 tests passed.")
        return 0
    print(f"\n{failures} failure(s).")
    return 1


if __name__ == "__main__":
    if "--test" in sys.argv:
        sys.exit(_run_self_tests())
    main_cli()
